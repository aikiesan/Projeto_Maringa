import sqlite3
import os
import openpyxl
from werkzeug.security import generate_password_hash

DB_PATH = os.path.join(os.path.dirname(__file__), 'database', 'maringa_project.db')

def init_and_seed_db():
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    if os.path.exists(DB_PATH):
        os.remove(DB_PATH)
        
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    # 1. Users Table
    cursor.execute('''
    CREATE TABLE users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT UNIQUE NOT NULL,
        name TEXT NOT NULL,
        password_hash TEXT NOT NULL,
        role TEXT NOT NULL, -- admin, pesquisador, visualizador
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )
    ''')
    
    # Insert initial users with hashed passwords
    users_data = [
        ('admin', 'Administrador LGPD', generate_password_hash('Maringa2026!'), 'admin'),
        ('pesquisador', 'Pesquisador Consultoria', generate_password_hash('Maringa2026!'), 'pesquisador'),
        ('visitante', 'Visualizador / Stakeholder', generate_password_hash('Maringa2026!'), 'visualizador')
    ]
    cursor.executemany('INSERT INTO users (username, name, password_hash, role) VALUES (?, ?, ?, ?)', users_data)
    
    # 2. Organizations Table (34 Orgs)
    cursor.execute('''
    CREATE TABLE organizations (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        code TEXT UNIQUE NOT NULL,
        name TEXT NOT NULL,
        acronym TEXT NOT NULL,
        sphere TEXT NOT NULL,
        nature TEXT NOT NULL,
        group_type TEXT NOT NULL,
        ccfla_main TEXT NOT NULL,
        ccfla_secondary TEXT,
        justification TEXT NOT NULL,
        contact_status TEXT NOT NULL,
        reference_materials TEXT
    )
    ''')
    
    # Load orgs from generated excel base
    wb = openpyxl.load_workbook('P3_Base_Mapeamento_Atores_Maringa.xlsx', data_only=True)
    ws_orgs = wb['Organizações (34)']
    orgs_to_insert = []
    for r in range(5, ws_orgs.max_row+1):
        c_code = ws_orgs.cell(r, 1).value
        c_name = ws_orgs.cell(r, 2).value
        c_acro = ws_orgs.cell(r, 3).value
        if not c_code or not c_name: continue
        c_sph = ws_orgs.cell(r, 4).value
        c_nat = ws_orgs.cell(r, 5).value
        c_grp = ws_orgs.cell(r, 6).value
        c_ccp = ws_orgs.cell(r, 7).value
        c_ccs = ws_orgs.cell(r, 8).value
        c_jus = ws_orgs.cell(r, 9).value
        c_sta = ws_orgs.cell(r, 10).value
        c_ref = ws_orgs.cell(r, 11).value
        orgs_to_insert.append((c_code, c_name, c_acro, c_sph, c_nat, c_grp, c_ccp, c_ccs, c_jus, c_sta, c_ref))
        
    cursor.executemany('''
    INSERT INTO organizations (code, name, acronym, sphere, nature, group_type, ccfla_main, ccfla_secondary, justification, contact_status, reference_materials)
    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', orgs_to_insert)
    
    # 3. People and Contacts Table
    cursor.execute('''
    CREATE TABLE people_contacts (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        code TEXT NOT NULL,
        contact_status TEXT NOT NULL,
        name TEXT NOT NULL,
        organization TEXT NOT NULL,
        acronym TEXT NOT NULL,
        role TEXT,
        group_type TEXT NOT NULL,
        email TEXT,
        phone TEXT,
        source TEXT,
        is_top_priority INTEGER DEFAULT 0
    )
    ''')
    
    ws_pess = wb['Pessoas e Contatos']
    pess_to_insert = []
    for r in range(5, ws_pess.max_row+1):
        p_code = ws_pess.cell(r, 1).value
        p_stat = ws_pess.cell(r, 2).value
        p_name = ws_pess.cell(r, 3).value
        if not p_code or not p_name: continue
        p_org = ws_pess.cell(r, 4).value
        p_acro = ws_pess.cell(r, 5).value
        p_role = ws_pess.cell(r, 6).value
        p_grp = ws_pess.cell(r, 7).value
        p_mail = ws_pess.cell(r, 8).value
        p_tel = ws_pess.cell(r, 9).value
        p_src = ws_pess.cell(r, 10).value
        is_top = 1 if p_stat == 'CONFIRMADO' else 0
        pess_to_insert.append((p_code, p_stat, p_name, p_org, p_acro, p_role, p_grp, p_mail, p_tel, p_src, is_top))
        
    cursor.executemany('''
    INSERT INTO people_contacts (code, contact_status, name, organization, acronym, role, group_type, email, phone, source, is_top_priority)
    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', pess_to_insert)
    
    # 4. COMDEMA Yearly Stats Table
    cursor.execute('''
    CREATE TABLE comdema_yearly_stats (
        year INTEGER PRIMARY KEY,
        total INTEGER NOT NULL,
        public_count INTEGER NOT NULL,
        private_count INTEGER NOT NULL,
        academia_count INTEGER NOT NULL,
        soc_civil_count INTEGER NOT NULL
    )
    ''')
    stats_data = [
        (2021, 47, 21, 10, 6, 10),
        (2022, 50, 23, 10, 7, 10),
        (2023, 46, 22, 8, 4, 12),
        (2024, 53, 26, 11, 4, 12),
        (2025, 48, 22, 10, 4, 12),
        (2026, 27, 13, 5, 3, 6)
    ]
    cursor.executemany('INSERT INTO comdema_yearly_stats VALUES (?, ?, ?, ?, ?, ?)', stats_data)
    
    # 5. COMDEMA Members Roster
    cursor.execute('''
    CREATE TABLE comdema_members (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        group_represented TEXT NOT NULL,
        affiliation TEXT NOT NULL,
        years_present TEXT NOT NULL,
        is_six_years INTEGER DEFAULT 0
    )
    ''')
    ws_com = wb['COMDEMA']
    # Extract members from tab 3
    members_to_insert = []
    # Roster starts at row 21 in COMDEMA sheet
    for r in range(21, ws_com.max_row+1):
        m_name = ws_com.cell(r, 1).value
        m_grp = ws_com.cell(r, 2).value
        m_aff = ws_com.cell(r, 3).value
        m_yrs = ws_com.cell(r, 4).value
        if not m_name: continue
        is_6 = 1 if '2021, 2022, 2023, 2024, 2025, 2026' in str(m_yrs) else 0
        members_to_insert.append((m_name, m_grp, m_aff, str(m_yrs), is_6))
        
    cursor.executemany('''
    INSERT INTO comdema_members (name, group_represented, affiliation, years_present, is_six_years)
    VALUES (?, ?, ?, ?, ?)
    ''', members_to_insert)
    
    # 6. Encoded Interviews Table (LGPD Compliant)
    cursor.execute('''
    CREATE TABLE encoded_interviews (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        interview_code TEXT UNIQUE NOT NULL,
        actor_code TEXT NOT NULL,
        institution_type TEXT NOT NULL,
        sector_group TEXT NOT NULL,
        ccfla_dimension TEXT NOT NULL,
        interview_date TEXT NOT NULL,
        instrument TEXT NOT NULL,
        status TEXT NOT NULL,
        key_findings_coded TEXT NOT NULL
    )
    ''')
    
    interviews_encoded_data = [
        ('INT-01', 'ACTOR-01', 'Secretaria Municipal de Saúde', 'Público', 'D1 - Planejamento e Política Climática', '2026-03-10', 'Entrevista Estruturada', 'Concluída & Codificada', 
         'Ator reporta necessidade de integração entre a vigilância zoológica/epidemiológica e as ações de adaptação climática. Destaca impactos de eventos térmicos na saúde pública.'),
        ('INT-02', 'ACTOR-02', 'Coordenadoria de Proteção e Defesa Civil', 'Público', 'D1 - Planejamento e Política Climática', '2026-03-12', 'Questionário Técnico', 'Concluída & Codificada', 
         'Mapeamento de pontos críticos de alagamento e enxurradas urbanas. Ponto focal confirmou protocolos de emergência e integração de alertas com a SEMOB e SEINFRA.'),
        ('INT-03', 'ACTOR-03', 'Universidade Estadual de Maringá (HUEM/GEMA)', 'Academia', 'D3 - Dados e Inteligência Climática', '2026-03-15', 'Entrevista de Profundidade', 'Concluída & Codificada', 
         'Detecção de lacunas de dados microclimáticos e ilhas de calor. Proposta de parceria acadêmica para modelos de projeção climática de alta resolução.'),
        ('INT-04', 'ACTOR-04', 'OAB & Instituto BiodiverCidade', 'Sociedade Civil', 'D4 - Governança e Participação', '2026-03-18', 'Entrevista Qualitativa', 'Concluída & Codificada', 
         'Análise do Arcabouço Jurídico do FUNDEMA e instrumentos tributários verdes (IPTU Verde). Recomenda fortalecimento do controle social e assentos no COMDEMA.'),
        ('INT-05', 'ACTOR-05', 'Secretaria de Urbanismo e Habitação (SEURBH)', 'Público', 'D1 - Planejamento e Política Climática', '2026-03-20', 'Entrevista Estruturada', 'Concluída & Codificada', 
         'Projetos de habitação de interesse social resiliente e revisão do Plano Diretor. Destaque para regularização fundiária em áreas de vulnerabilidade socioambiental.'),
        ('INT-06', 'ACTOR-06', 'Instituto Ambiental de Maringá (IAM)', 'Público', 'D1 - Planejamento e Política Climática', '2026-03-22', 'Questionário Técnico', 'Concluída & Codificada', 
         'Gestão de unidades de conservação urbanas e licenciamento de empreendimentos. Relato de desafios no monitoramento de áreas de preservação permanente (APPs).'),
        ('INT-07', 'ACTOR-07', 'Secretaria de Limpeza Urbana (SELURB)', 'Público', 'D1 - Planejamento e Política Climática', '2026-03-24', 'Entrevista Estruturada', 'Concluída & Codificada', 
         'Plano de arborização urbana e gestão de resíduos sólidos. Identificação de rotas prioritárias para sombreamento e mitigação de ilhas de calor.'),
        ('INT-08', 'ACTOR-08', 'NAPI Agenda 2030 / COMDEMA / UEM', 'Academia', 'D4 - Governança e Participação', '2026-03-26', 'Entrevista de Profundidade', 'Concluída & Codificada', 
         'Alinhamento com os Objetivos de Desenvolvimento Sustentável (ODS 11 e 13). Recomenda governança multinível entre Município, UEM e COMDEMA.'),
        ('INT-09', 'ACTOR-09', 'Secretaria de Infraestrutura (SEINFRA)', 'Público', 'D2 - Financiamento e Capacidade Fiscal', '2026-03-28', 'Entrevista Estruturada', 'Concluída & Codificada', 
         'Carteira de projetos de obras estruturantes de drenagem e pavimentação permeável. Necessidade de captação externa e apoio na estruturação de financiamento.'),
        ('INT-10', 'ACTOR-10', 'Instituto de Pesquisa e Planejamento (IPPLAM)', 'Público', 'D3 - Dados e Inteligência Climática', '2026-04-02', 'Oficina Técnica', 'Concluída & Codificada', 
         'Integração de dados cartográficos e supervisão do projeto de financiamento climático. IPPLAM atua como articulador central do ecossistema municipal.'),
        ('INT-11', 'ACTOR-11', 'Secretaria de Governo (SEGOV)', 'Público', 'D4 - Governança e Participação', '2026-04-05', 'Entrevista Estruturada', 'Concluída & Codificada', 
         'Articulação política e priorização de projetos no PPA (Plano Plurianual). Confirmação de suporte institucional à agenda de financiamento climático.'),
        ('INT-12', 'ACTOR-12', 'CODEM / IDR-Paraná (EMATER)', 'Privado / Público', 'D2 - Financiamento e Capacidade Fiscal', '2026-04-08', 'Entrevista de Profundidade', 'Concluída & Codificada', 
         'Integração de cadeias produtivas sustentáveis, crédito rural verde e suporte a pequenos produtores periurbanos em técnicas de conservação de solo e água.')
    ]
    cursor.executemany('''
    INSERT INTO encoded_interviews (interview_code, actor_code, institution_type, sector_group, ccfla_dimension, interview_date, instrument, status, key_findings_coded)
    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', interviews_encoded_data)
    
    # 7. Audit Logs Table (LGPD compliance)
    cursor.execute('''
    CREATE TABLE audit_logs (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        user_id INTEGER,
        username TEXT NOT NULL,
        action TEXT NOT NULL,
        details TEXT NOT NULL
    )
    ''')
    cursor.execute('INSERT INTO audit_logs (user_id, username, action, details) VALUES (1, "admin", "SISTEMA_INICIADO", "Banco de dados inicializado com criptografia e regras de privacidade LGPD.")')
    
    conn.commit()
    conn.close()
    print('Banco de dados SQLite maringa_project.db inicializado e povoado com sucesso!')

if __name__ == '__main__':
    init_and_seed_db()
