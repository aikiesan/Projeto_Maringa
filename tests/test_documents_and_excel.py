import unittest
import openpyxl
import docx
import json
import os

class TestDocumentsAndExcel(unittest.TestCase):

    def setUp(self):
        self.root_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
        self.excel_path = os.path.join(self.root_dir, 'P3_Base_Mapeamento_Atores_Maringa.xlsx')
        self.word_path = os.path.join(self.root_dir, 'P3_Relatorio_Mapeamento_Atores_Maringa.docx')

    def test_01_excel_sheets_structure(self):
        """Verifica se a planilha Excel contém as 4 abas exatas exigidas e 34 organizações."""
        self.assertTrue(os.path.exists(self.excel_path), "Planilha P3_Base_Mapeamento_Atores_Maringa.xlsx não encontrada.")
        wb = openpyxl.load_workbook(self.excel_path, data_only=True)
        
        expected_sheets = ['Organizações (34)', 'Pessoas e Contatos', 'COMDEMA', 'Painel']
        self.assertEqual(wb.sheetnames, expected_sheets, f"As abas devem ser exatamente: {expected_sheets}")

        # Check 34 Orgs count
        ws_orgs = wb['Organizações (34)']
        org_count = 0
        for r in range(5, ws_orgs.max_row+1):
            if ws_orgs.cell(r, 1).value: org_count += 1
        self.assertEqual(org_count, 34, "A aba Organizações (34) deve conter exatamente 34 entidades.")

        # Check confirmed contacts on top
        ws_pess = wb['Pessoas e Contatos']
        top_status = [ws_pess.cell(r, 2).value for r in range(5, 10)]
        self.assertTrue(all(s == 'CONFIRMADO' for s in top_status), "Os primeiros 5 contatos da aba Pessoas e Contatos devem ter status CONFIRMADO.")

    def test_02_word_report_sections(self):
        """Verifica se o Relatório Word contém todos os tópicos e seções exigidos."""
        self.assertTrue(os.path.exists(self.word_path), "Relatório P3_Relatorio_Mapeamento_Atores_Maringa.docx não encontrado.")
        doc = docx.Document(self.word_path)
        
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        doc_text = " ".join(paragraphs)

        required_keywords = [
            "PRODUTO 3", "Mapeamento Institucional", "CCFLA/CEPAL",
            "1. Apresentação", "2. Objetivo do Mapeamento", "3. Metodologia e Fontes",
            "4. Critérios de Inclusão", "5. Mapeamento dos Atores Organizado por Bloco",
            "6. Governança Ambiental: Análise do COMDEMA", "7. Lacunas Identificadas", "8. Próximos Passos",
            "SEPLAN", "Fazenda"
        ]

        for kw in required_keywords:
            self.assertIn(kw, doc_text, f"O relatório Word deve conter a palavra-chave/seção: '{kw}'")

    def test_03_static_json_validity(self):
        """Verifica se os arquivos JSON na pasta data/ e docs/data/ são sintaticamente válidos."""
        for sub in ['data', os.path.join('docs', 'data')]:
            json_dir = os.path.join(self.root_dir, sub)
            for filename in ['overview.json', 'organizations.json', 'contacts.json', 'comdema.json', 'interviews.json']:
                file_path = os.path.join(json_dir, filename)
                self.assertTrue(os.path.exists(file_path), f"Arquivo JSON {file_path} não encontrado.")
                with open(file_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    self.assertIsNotNone(data, f"Arquivo {file_path} não é um JSON válido.")

if __name__ == '__main__':
    unittest.main()
